"""
Sube casos de prueba (Excel, formato `casos de prueba/plantilla_base.xlsx`) como
Test Cases de QMetry (QTM4J), vía la API interna `/rest/qtm4j/ui/latest/testcases`.

Autenticación: Basic Auth con JIRA_USERNAME/JIRA_API_TOKEN (no se usa Open API Key de
QMetry; la creación de Test Cases solo se validó contra este endpoint con Basic Auth).

Uso:
    python jira_uploader.py "casos de prueba/suite_automatizacion_everest.xlsx"
"""
import os
import sys

import openpyxl
import requests
from dotenv import load_dotenv

COLUMNAS_ESPERADAS = [
    "Issue ID", "Tipo de test", "Resumen", "Descripcion", "Escenario",
    "Resultado Final", "Accion", "Datos", "Resultado Esperado",
]
FOLDER_ID_DEFAULT = -1  # sin carpeta, raíz del proyecto (decisión del usuario)
PRIORITY_ID_DEFAULT = 1906  # "High" (confirmado vía GET .../priorities del proyecto 79906)
STATUS_ID_DEFAULT = 4290  # "To Do" (confirmado vía GET .../testcase-statuses del proyecto 79906)


def _cargar_configuracion():
    load_dotenv()
    faltantes = [
        var for var in ("JIRA_URL", "JIRA_USERNAME", "JIRA_API_TOKEN", "QMETRY_PROJECT_ID")
        if not os.getenv(var)
    ]
    if faltantes:
        raise RuntimeError(f"Faltan variables en .env: {', '.join(faltantes)}")
    return {
        "jira_url": os.environ["JIRA_URL"].rstrip("/"),
        "username": os.environ["JIRA_USERNAME"],
        "token": os.environ["JIRA_API_TOKEN"],
        "project_id": int(os.environ["QMETRY_PROJECT_ID"]),
    }


def _indice_columnas(hoja):
    cabeceras = {celda.value: celda.column for celda in hoja[1] if celda.value}
    faltantes = [c for c in COLUMNAS_ESPERADAS if c not in cabeceras]
    if faltantes:
        raise ValueError(f"El Excel no tiene las columnas esperadas: {', '.join(faltantes)}")
    return cabeceras


def _payload_test_case(fila, project_id):
    return {
        "summary": fila["Resumen"] or "",
        "description": fila["Descripcion"] or "",
        "precondition": fila["Escenario"] or "",
        "folderId": FOLDER_ID_DEFAULT,
        "projectId": project_id,
        "priority": PRIORITY_ID_DEFAULT,
        "status": STATUS_ID_DEFAULT,
        "steps": [{
            "stepDetails": fila["Accion"] or "",
            "testData": fila["Datos"] or "",
            "expectedResult": fila["Resultado Esperado"] or "",
            "isChecked": False,
            "isExpanded": True,
        }],
    }


def _crear_test_case(sesion, config, payload):
    url = f"{config['jira_url']}/rest/qtm4j/ui/latest/testcases"
    respuesta = sesion.post(url, json=payload, timeout=30)
    if not respuesta.ok:
        raise RuntimeError(f"HTTP {respuesta.status_code}: {respuesta.text[:300]}")
    cuerpo = respuesta.json()
    return cuerpo["key"]


def subir_casos_a_qmetry(ruta_excel):
    config = _cargar_configuracion()
    wb = openpyxl.load_workbook(ruta_excel)
    hoja = wb.active
    cols = _indice_columnas(hoja)

    sesion = requests.Session()
    sesion.auth = (config["username"], config["token"])

    creados, fallidos = {}, []

    for num_fila in range(2, hoja.max_row + 1):
        resumen = hoja.cell(row=num_fila, column=cols["Resumen"]).value
        if not resumen:
            continue  # fila vacía

        fila = {nombre: hoja.cell(row=num_fila, column=col).value for nombre, col in cols.items()}
        payload = _payload_test_case(fila, config["project_id"])
        try:
            key = _crear_test_case(sesion, config, payload)
        except Exception as error:
            fallidos.append((num_fila, resumen, str(error)))
            continue

        creados[num_fila] = key

    return {"creados": creados, "fallidos": fallidos}


def _imprimir_resumen(resultado):
    print(f"\nCreados : {len(resultado['creados'])}")
    for fila, key in resultado["creados"].items():
        print(f"  fila {fila} -> {key}")
    print(f"Fallidos: {len(resultado['fallidos'])}")
    for fila, resumen, error in resultado["fallidos"]:
        print(f"  fila {fila} [{resumen}] -> {error}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python jira_uploader.py <ruta_al_excel>")
        sys.exit(1)
    resultado = subir_casos_a_qmetry(sys.argv[1])
    _imprimir_resumen(resultado)
    sys.exit(1 if resultado["fallidos"] else 0)
